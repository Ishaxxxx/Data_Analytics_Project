import pandas as pd
import tkinter as tk
from tkinter import filedialog, simpledialog
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from win32com.client import Dispatch
import win32com.client.gencache
from datetime import datetime
import os
import sys
from colorama import init, Fore, Style
import random
import subprocess
from typing import Optional, Tuple

init(autoreset=True)


BLUE_HEADER = "0070C0"
PINK_SLA = "FFC7CE"
FALLBACK_REMARKS = [
    "visit pending", "customer not present", "call closed by 4 pm",
    "part pending", "custom remarks"
]
COLUMN_WIDTHS = [18, 8, 22, 50, 15, 20, 35, 20, 22, 30, 30]  # Added one more column for Remarks


def log(message: str, level: str = "info") -> None:
    """Enhanced logging function with colors and timestamps."""
    timestamp = datetime.now().strftime("%H:%M:%S")
    levels = {
        "info": Fore.CYAN,
        "success": Fore.GREEN,
        "error": Fore.RED,
        "step": Fore.YELLOW,
        "warning": Fore.MAGENTA
    }
    color = levels.get(level, Fore.WHITE)
    print(f"{Fore.WHITE}[{timestamp}] {color}[{level.upper()}] {Style.RESET_ALL}{message}")


def create_tk_root() -> tk.Tk:
    """Create and configure a Tk root window."""
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    root.focus_force()
    return root

def select_file(title: str, filetypes: list) -> Optional[str]:
    """Generic file selection function."""
    root = create_tk_root()
    try:
        file_path = filedialog.askopenfilename(title=title, filetypes=filetypes, parent=root)
        return file_path
    finally:
        root.destroy()

def select_csv_file() -> Optional[str]:
    """Select a CSV file."""
    return select_file("Select CSV File", [("CSV Files", "*.csv")])

def select_output_folder() -> Optional[str]:
    """Select an output folder."""
    root = create_tk_root()
    try:
        folder_path = filedialog.askdirectory(title="Select Output Folder", parent=root)
        return folder_path
    finally:
        root.destroy()

def select_lookup_file() -> Optional[str]:
    """Select an Excel file for VLOOKUP."""
    return select_file(
        "Select Lookup Excel File",
        [("Excel Files", "*.xlsx"), ("Excel Files", "*.xls")]
    )


def read_lookup_file(lookup_path: str) -> Optional[pd.DataFrame]:
    """Read the lookup Excel file with multiple engine fallbacks."""
    engines = ['openpyxl', 'xlrd']
    for engine in engines:
        try:
            log(f"Trying to read lookup file with {engine} engine...", "info")
            lookup_df = pd.read_excel(lookup_path, sheet_name='Sheet1', engine=engine)
            log(f"Successfully read lookup file with {engine}", "success")
            return lookup_df
        except Exception as e:
            log(f"Failed to read with {engine}: {str(e)}", "warning")
    
    log("All engines failed to read the lookup file", "error")
    return None

def validate_columns(df: pd.DataFrame, required_columns: list, source: str) -> bool:
    """Validate that required columns exist in the dataframe."""
    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        log(f"Missing columns in {source}: {missing}", "error")
        return False
    return True

def apply_vlookup(df: pd.DataFrame, lookup_path: str) -> pd.DataFrame:
    """Apply VLOOKUP functionality to merge data from lookup file."""
    log("Starting VLOOKUP process...", "step")
    
    lookup_df = read_lookup_file(lookup_path)
    if lookup_df is None:
        return df
    
    log(f"Lookup file columns: {list(lookup_df.columns)}", "debug")
    
    key_column, value_column = determine_lookup_columns(df, lookup_df)
    if not key_column or not value_column:
        return df
    
    df = clean_key_column(df, key_column)
    lookup_df = clean_key_column(lookup_df, key_column)
    
    log_sample_data(df, lookup_df, key_column)
    
    merged = perform_merge(df, lookup_df, key_column, value_column)
    
    if merged is not None:
        df = process_merge_results(merged, value_column)
    
    return df

def determine_lookup_columns(df: pd.DataFrame, lookup_df: pd.DataFrame) -> Tuple[str, str]:
    """Determine the key and value columns for VLOOKUP."""
    default_key = "Case Number"
    default_value = "Remarks"
    
    if default_key not in lookup_df.columns:
        key_column = prompt_user_column(
            "Enter the COMMON KEY COLUMN (like Case Number):",
            list(lookup_df.columns)
        )
        if not key_column:
            return "", ""
    else:
        key_column = default_key
    
    if key_column not in df.columns:
        log(f"Key column '{key_column}' not found in main CSV", "error")
        return "", ""
    
    if default_value not in lookup_df.columns:
        value_column = prompt_user_column(
            "Enter the LOOKUP RETURN COLUMN (like Remarks):",
            list(lookup_df.columns)
        )
        if not value_column:
            return key_column, ""
    else:
        value_column = default_value
    
    return key_column, value_column

def prompt_user_column(prompt: str, columns: list) -> str:
    """Prompt user to select a column from a list."""
    root = create_tk_root()
    try:
        column = simpledialog.askstring(
            "Input", 
            f"{prompt}\nAvailable columns:\n{', '.join(columns)}\n\nColumn name:"
        )
        return column if column in columns else ""
    finally:
        root.destroy()

def clean_key_column(df: pd.DataFrame, column: str) -> pd.DataFrame:
    """Clean the key column for matching."""
    df[column] = df[column].astype(str).str.replace('.0', '', regex=False)
    return df

def log_sample_data(df: pd.DataFrame, lookup_df: pd.DataFrame, key_column: str) -> None:
    """Log sample data for verification."""
    log(f"Sample CSV data - {key_column}: {df[key_column].head(3).tolist()}", "debug")
    log(f"Sample lookup data - {key_column}: {lookup_df[key_column].head(3).tolist()}", "debug")
    
    common_cases = set(df[key_column]).intersection(set(lookup_df[key_column]))
    log(f"Common cases found: {len(common_cases)}", "info")
    if common_cases:
        log(f"Sample common cases: {list(common_cases)[:5]}", "debug")

def perform_merge(df: pd.DataFrame, lookup_df: pd.DataFrame, 
                 key_column: str, value_column: str) -> Optional[pd.DataFrame]:
    """Perform the merge operation between dataframes."""
    try:
        merged = df.merge(
            lookup_df[[key_column, value_column]], 
            on=key_column, 
            how='left'
        )
        log(f"Merged successfully. New columns: {list(merged.columns)}", "debug")
        return merged
    except Exception as e:
        log(f"Merge failed: {str(e)}", "error")
        return None

def process_merge_results(merged: pd.DataFrame, value_column: str) -> pd.DataFrame:
    """Process the results of the merge operation."""
    
    original_remarks_col = 'Remarks_x' if 'Remarks_x' in merged.columns else 'Remarks'
    lookup_remarks_col = 'Remarks_y' if 'Remarks_y' in merged.columns else value_column
    
    
    if lookup_remarks_col in merged.columns:
        merged[lookup_remarks_col] = merged[lookup_remarks_col].fillna('')
    
    merged['Remarks'] = ''
    
    vlookup_overwrites = 0
    original_kept = 0
    
    for idx, row in merged.iterrows():
        
        vlookup_result = ''
        if lookup_remarks_col in merged.columns:
            vlookup_result = str(row[lookup_remarks_col]).strip()
        
        # Only populate the Remarks column if we have valid VLOOKUP data
        # Don't use original_remark since the Remarks column should only have VLOOKUP data
        if vlookup_result and vlookup_result.lower() not in ['nan', '', 'none']:
            merged.at[idx, 'Remarks'] = vlookup_result
            vlookup_overwrites += 1
        else:
            # Keep the Remarks column empty if no VLOOKUP data
            merged.at[idx, 'Remarks'] = ''
            original_kept += 1
    
    cols_to_drop = []
    if 'Remarks_x' in merged.columns:
        cols_to_drop.append('Remarks_x')
    if 'Remarks_y' in merged.columns:
        cols_to_drop.append('Remarks_y')
    if value_column != 'Remarks' and value_column in merged.columns and value_column not in cols_to_drop:
        cols_to_drop.append(value_column)
    
    if cols_to_drop:
        merged = merged.drop(cols_to_drop, axis=1)
    
    log(f"VLOOKUP Results: {vlookup_overwrites} rows populated with lookup data, {original_kept} rows left empty", "success")
    
    return merged


def read_csv_with_fallback(csv_path: str) -> Optional[pd.DataFrame]:
    """Read CSV with encoding fallback."""
    encodings = ['utf-8', 'latin1', 'utf-16']
    for encoding in encodings:
        try:
            log(f"Trying to read CSV with {encoding} encoding...", "debug")
            return pd.read_csv(csv_path, encoding=encoding, low_memory=False)
        except UnicodeDecodeError:
            continue
    log("Failed to read CSV with all attempted encodings", "error")
    return None

def process_csv(csv_path: str, output_path: str, 
               do_vlookup: bool = False, lookup_path: Optional[str] = None) -> bool:
    """Main CSV processing function."""
    log("Starting CSV processing...", "step")
    
    df = read_csv_with_fallback(csv_path)
    if df is None:
        return False
    
    df = prepare_dataframe(df)
    
    if do_vlookup and lookup_path:
        df = apply_vlookup(df, lookup_path)
    
    if not save_to_excel(df, output_path):
        return False
    
    create_pivot_table(output_path)
    
    return True

def prepare_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Prepare and clean the dataframe."""
    df['Created Date'] = pd.to_datetime(df['Created Date'], dayfirst=True, errors='coerce')
    df['SLA'] = (datetime.now() - df['Created Date']).dt.days.fillna(0).astype(int)

    if 'LineItem Status' not in df.columns:
        df['LineItem Status'] = 'New'
    else:
        df['LineItem Status'] = df['LineItem Status'].fillna('New').astype(str)
    
    # Select and order columns - keep Technician Remarks as it is
    selected_columns = [
        'Case Number', 'SLA', 'Customer Name', 'Street',
        'Zip/Postal Code', 'Customer Complaint', 'Product Description',
        'LineItem Status', 'Technician Name'
    ]
    
    # Add Technician Remarks if it exists (keep original column)
    if 'Technician Remarks' in df.columns:
        selected_columns.append('Technician Remarks')
        log("Keeping Technician Remarks column as it is", "success")
    
    # Always add a new empty Remarks column (separate from Technician Remarks)
    df['Remarks'] = ''  # Start with empty column
    selected_columns.append('Remarks')
    log("Added new empty Remarks column for VLOOKUP results", "success")
    
    return df[selected_columns]


def save_to_excel(df: pd.DataFrame, excel_path: str) -> bool:
    """Save dataframe to Excel with formatting."""
    try:
        log("Creating Excel workbook...", "step")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
        
        apply_excel_styling(ws, df)
        
        wb.save(excel_path)
        log(f"Excel file saved to {excel_path}", "success")
        return True
    except Exception as e:
        log(f"Error saving Excel file: {str(e)}", "error")
        return False

def apply_excel_styling(ws, df: pd.DataFrame) -> None:

    header_fill = PatternFill(start_color=BLUE_HEADER, end_color=BLUE_HEADER, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'),
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    pink_fill = PatternFill(start_color=PINK_SLA, end_color=PINK_SLA, fill_type="solid")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = align
    
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.fill = pink_fill
    
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = align
    
    for i, width in enumerate(COLUMN_WIDTHS, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    ws.auto_filter.ref = ws.dimensions


def create_pivot_table(excel_path: str) -> None:
    """Create pivot table in the Excel file."""
    log("Attempting to create pivot table...", "step")
    
    kill_excel_processes()
    
    excel = None
    wb = None
    
    try:
        excel = initialize_excel_application()
        if not excel:
            return
        
        abs_path = os.path.abspath(excel_path)
        wb = excel.Workbooks.Open(abs_path)
        ws = wb.Sheets("Sheet1")
        
        apply_default_sheet_settings(ws)
        
        last_row = ws.UsedRange.Rows.Count
        data_range = ws.Range(f"A1:K{last_row}")  # Updated to K since we now have 11 columns
        
        ws_pivot = wb.Sheets.Add()
        ws_pivot.Name = "Pivot_View"
        
        pivot_cache = wb.PivotCaches().Create(1, data_range) 
        pivot_table = pivot_cache.CreatePivotTable(ws_pivot.Range("B4"), "SLA_Pivot")
        
        configure_pivot_fields(pivot_table)
        
        wb.Save()
        log("Pivot table created successfully", "success")
        
    except Exception as e:
        log(f"Error creating pivot table: {str(e)}", "error")
    finally:
        if wb:
            try:
                wb.Close(True)
            except:
                pass
        if excel:
            try:
                excel.Quit()
            except:
                pass

def kill_excel_processes() -> None:
    """Kill any running Excel processes."""
    try:
        subprocess.run("taskkill /f /im excel.exe", shell=True, 
                      stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except:
        pass

def initialize_excel_application():
    """Initialize Excel application with multiple fallback methods."""
    methods = [
        lambda: win32com.client.gencache.EnsureDispatch('Excel.Application'),
        lambda: win32com.client.Dispatch('Excel.Application'),
        lambda: win32com.client.dynamic.Dispatch('Excel.Application')
    ]
    
    for method in methods:
        try:
            excel = method()
            excel.Visible = False
            excel.DisplayAlerts = False
            return excel
        except Exception as e:
            log(f"Excel initialization attempt failed: {str(e)}", "debug")
    
    log("All Excel initialization methods failed", "error")
    return None

def configure_pivot_fields(pivot_table) -> None:
    """Configure pivot table fields and layout."""
    try:
        pf_row = pivot_table.PivotFields("Technician Name")
        pf_row.Orientation = 1 
        
        pf_col = pivot_table.PivotFields("SLA")
        pf_col.Orientation = 2  
        
        pf_filter = pivot_table.PivotFields("LineItem Status")
        pf_filter.Orientation = 3  
        
        try:
            pf_filter.CurrentPage = "New"
        except:
            log("Could not set default filter to 'New'", "warning")
        
        data_field = pivot_table.AddDataField(
            pivot_table.PivotFields("Case Number"), 
            "Count of Case Number", 
            -4112 
        )
        
        try:
            pf_row.AutoSort(2, "Count of Case Number") 
        except:
            log("Could not apply auto-sort to pivot table", "warning")
            
    except Exception as e:
        log(f"Error configuring pivot fields: {str(e)}", "error")

def apply_default_sheet_settings(ws) -> None:
    """Apply default sorting and filtering to the main data sheet while keeping all filter options available."""
    try:
        last_row = ws.UsedRange.Rows.Count
        last_col = ws.UsedRange.Columns.Count
        data_range = ws.Range(ws.Cells(1, 1), ws.Cells(last_row, last_col))
        
        data_range.AutoFilter(Field=1)  
        
        headers = [str(ws.Cells(1, col).Value).strip() for col in range(1, last_col+1)]
        try:
            sla_col = headers.index("SLA") + 1
        except ValueError:
            sla_col = None
            
        try:
            lineitem_col = headers.index("LineItem Status") + 1
        except ValueError:
            lineitem_col = None
        
        if sla_col:
            ws.Sort.SortFields.Clear()
            ws.Sort.SortFields.Add(
                Key=ws.Cells(1, sla_col),
                SortOn=0,  
                Order=2    
            )
            ws.Sort.SetRange(data_range)
            ws.Sort.Header = 1  
            ws.Sort.Apply()
            log("Applied default SLA sorting (descending)", "success")
        
        if lineitem_col:
            data_range.AutoFilter(
                Field=lineitem_col,
                Criteria1="New"
            )
            log("Applied default LineItem Status filter (New)", "success")
        
    except Exception as e:
        log(f"Error applying sheet settings: {str(e)}", "error")
        try:
            ws.UsedRange.AutoFilter()
        except:
            pass


def main() -> None:
    """Main entry point for the script."""
    print(Fore.GREEN + "\n===============================")
    print(Fore.CYAN + "🚀 CSV Formatter with Excel VLOOKUP & Pivot Table")
    print(Fore.GREEN + "===============================\n")
    
    try:
        csv_path = select_csv_file()
        if not csv_path:
            log("No CSV file selected. Exiting...", "error")
            return
        
        output_folder = select_output_folder()
        if not output_folder:
            log("No output folder selected. Exiting...", "error")
            return
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Output_{timestamp}.xlsx"
        excel_path = os.path.join(output_folder, filename)
        
        do_vlookup = input("Apply VLOOKUP? (Y/N): ").strip().lower() == 'y'
        lookup_file = select_lookup_file() if do_vlookup else None
        
        if do_vlookup and not lookup_file:
            log("No lookup file selected. Skipping VLOOKUP...", "warning")
            do_vlookup = False
        
        if process_csv(csv_path, excel_path, do_vlookup, lookup_file):
            try:
                os.startfile(excel_path)
                log("🎉 Processing complete! File opened.", "success")
            except:
                log("Processing complete! Unable to automatically open the file.", "success")
    
    except Exception as e:
        log(f"Unexpected error: {str(e)}", "error")
        sys.exit(1)

if __name__ == "__main__":
    main()